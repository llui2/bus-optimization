# script/line_by_line_SA.py

import pandas as pd
import numpy as np
import os
from datetime import datetime
from openpyxl import load_workbook # type: ignore
from informal import euclidean_distance
import networkx as nx


class DataHandler:
    def __init__(self, excel_path="data/comparatives_imatges.xlsx", sheet_name="SA"):
        self.excel_path = excel_path
        self.sheet_name = sheet_name

    def compute_metrics(self, line_nodes, obj, pos, node_to_bus_index, D, G, d):
        final_energy = obj[0]

        travel_distances = []
        population_total = 0

        for i in range(len(line_nodes)):
            for j in range(len(line_nodes)):
                if i != j:
                    src = line_nodes[i]
                    dst = line_nodes[j]
                    index_src = node_to_bus_index[src]
                    index_dst = node_to_bus_index[dst]
                    demand = D[index_src, index_dst]
                    if demand > 0:
                        try:
                            path = nx.shortest_path(G, src, dst, weight="cost")
                            cost = sum(G.edges[path[k], path[k+1]]['cost'] for k in range(len(path)-1))
                            travel_distances.append(cost)
                            population_total += demand
                        except nx.NetworkXNoPath:
                            pass

        avg_travel_distance = np.round(np.mean(travel_distances), 2) if travel_distances else 0

        center_x, center_y = d / 2, d / 2
        radio = d / 6
        covered_central = [node for node in line_nodes if euclidean_distance(pos[node], (center_x, center_y)) <= radio]
        central_coverage_pct = np.round(100 * len(covered_central) / len(line_nodes), 1)

        # Calcular l'iteració de convergència a partir de data/energy.csv
        try:
            energy_df = pd.read_csv("data/energy.csv")
            energy_series = energy_df['e0']
            diffs = energy_series.diff().abs()
            converged_at = diffs.gt(1).to_numpy().nonzero()[0]
            converged_iter = int(converged_at[-1]) if len(converged_at) > 0 else 0
        except Exception:
            converged_iter = -1  # valor per defecte si hi ha error

        return final_energy, avg_travel_distance, population_total, central_coverage_pct, converged_iter


    def export_to_excel(self, metrics_dict):
        metrics_df = pd.DataFrame([metrics_dict])

        if not os.path.exists(self.excel_path):
            with pd.ExcelWriter(self.excel_path, engine="openpyxl") as writer:
                metrics_df.to_excel(writer, sheet_name=self.sheet_name, index=False)
            return

        #book = load_workbook(self.excel_path)
        #writer = pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
        #writer.sheets = {ws.title: ws for ws in book.worksheets}

        #start_row = writer.sheets[self.sheet_name].max_row + 1

        writer = pd.ExcelWriter(self.excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay')

        start_row = writer.sheets[self.sheet_name].max_row + 1

        metrics_df.to_excel(writer, sheet_name=self.sheet_name, index=False, header=False, startrow=start_row)
        writer.close()

    def prepare_metrics_dict(self, final_energy, avg_travel_distance, population_total,
                            central_coverage_pct, seed, converged_iter, L, T_0, lambdaa, max_iter):

        return {
            "final_energy": final_energy,
            "avg_travel_distance": avg_travel_distance,
            "population_transported": np.round(population_total, 2),
            "central_coverage_pct": central_coverage_pct,
            "converged_iter": converged_iter,
            "datetime": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "seed": seed,
            "L": L,
            "T_0": T_0,
            "lambdaa": lambdaa,
            "max_iter": max_iter
        }

